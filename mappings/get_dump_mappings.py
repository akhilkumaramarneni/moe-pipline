import pickle
import openpyxl
import json
import os

current_phases = [1, 2, 3, 4, 5, 6, 7, 8]


# below code to generate pahes to detector mappings for one region
# later add same for other region, if same excel type can reuse it - FIU and Gainsville

def advance_det_mappings(region):
    print("generating advance detector mappings for", region)
    advance_detectors={}
    workbook = openpyxl.load_workbook("testmoe.xlsx")
    worksheet = workbook.active
    # R-19(distance), O-14(detectors)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[19]>15:
            if row[1] not in advance_detectors:
                advance_detectors[row[1]]= set()
            advance_detectors[row[1]].add(row[14])
            
    advance_detectors = {key: list(values) for key, values in advance_detectors.items()}

    # for testing - TODO : remove later
    advance_detectors["916"]= [ 9, 10, 13, 14, 25, 51, 52, 35, 41, 42, 45, 46, 57, 59 ]
    advance_detectors["1430"] = [ 9, 10, 13, 14, 25, 51, 52, 35, 41, 42, 45, 46, 57, 59 ]

    with open(region + "/" + "advance_detector_mappings.pickle", "wb") as pickle_file:
        pickle.dump(advance_detectors, pickle_file)
    print("generated advance detector mappings done for ", region)
    
    # print(advance_detectors)

def phase_to_detector_mappings(region):
    print("generating phase to detector mappings for", region)
    # Q-16(phase), p-15(channel)
    all_maps = {}
    workbook = openpyxl.load_workbook("testmoe.xlsx")
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Process the data in the row
        if row[1] not in all_maps:
            all_maps[row[1]] = {i: set() for i in range(1, 9)}
        if row[16] not in current_phases:
            continue
        all_maps[row[1]][row[16]].add(row[15])

    # just format change
    for intersection in all_maps.keys():
        for phase in all_maps[intersection].keys():
            all_maps[intersection][phase] = list(all_maps[intersection][phase])

    # can use it or we can dump object to file and use it
    # region_mappings = {}
    # region_mappings[region] = all_maps
    region_mappings = all_maps
    # print(region_mappings)
    # Dump data into a JSON file

    # for testing - TODO : remove later
    all_maps["916"] = {
        1: [],
        2: [5, 6, 7, 9, 10, 13, 14],
        3: [],
        4: [21, 49, 25, 51, 52],
        5: [33, 35],
        6: [37, 38, 39, 41, 42, 45, 46],
        7: [],
        8: [53, 57, 59],
    }

    all_maps["1430"] = {
        1: [],
        2: [5, 6, 7, 9, 10, 13, 14],
        3: [],
        4: [21, 49, 25, 51, 52],
        5: [33, 35],
        6: [37, 38, 39, 41, 42, 45, 46],
        7: [],
        8: [53, 57, 59],
    }
    
    if not os.path.exists(region):
        os.makedirs(region)

    with open(region + "/" + "phase_to_detector_mappings.pickle", "wb") as pickle_file:
        pickle.dump(region_mappings, pickle_file)
    print("generated phase to detector mappings done for ", region)
    return all_maps

def mapped_phases(region, phase_to_detector_map):
    print("generating map phases for", region)
    mappings = {}
    for key, values in phase_to_detector_map.items():
        non_empty_phases = [k for k, v in values.items() if v]
        mappings[key] = non_empty_phases

    with open(region + "/" + "mapped_phases.pickle", "wb") as pickle_file:
        pickle.dump(mappings, pickle_file)

    print("generated map phases for", region)
    # print(mappings)
     
    

if __name__ == "__main__":
    region = "Orlando"
    phase_to_detector_map = phase_to_detector_mappings(region)
    mapped_phases(region,phase_to_detector_map)
    advance_det_mappings(region)
